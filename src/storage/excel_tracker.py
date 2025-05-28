import pandas as pd
import os
import uuid
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

class ExcelContentTracker:
    def __init__(self, excel_path="data/content_tracker.xlsx"):
        self.excel_path = excel_path
        self.ensure_excel_exists()
        
    def ensure_excel_exists(self):
        """Создает Excel-файл с необходимой структурой, если он не существует"""
        if not os.path.exists(self.excel_path):
            # Создаем директорию, если она не существует
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            
            # Создаем DataFrame для каждого листа
            content_df = pd.DataFrame(columns=[
                'content_id', 'article_id', 'title', 'source_url', 'category',
                'content_type', 'language', 'content_markdown', 'creation_date',
                'status', 'scheduled_date', 'scheduled_time', 'platform',
                'published_date', 'published_url', 'engagement_stats',
                'tags', 'dubskiy_rating', 'notes'
            ])
            
            articles_df = pd.DataFrame(columns=[
                'article_id', 'title', 'source_url', 'category',
                'publication_date', 'company_name', 'funding_amount',
                'article_content', 'processing_date', 'processing_status'
            ])
            
            schedule_df = pd.DataFrame(columns=[
                'schedule_id', 'content_id', 'platform', 'scheduled_date',
                'scheduled_time', 'timezone', 'posting_status', 'priority',
                'campaign_id', 'posting_account'
            ])
            
            metadata_df = pd.DataFrame({
                'key': ['last_update', 'version', 'api_keys', 'default_settings', 'platform_settings'],
                'value': [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '1.0',
                    '{}',
                    '{"default_time": "10:00", "default_timezone": "UTC+3"}',
                    '{}'
                ]
            })
            
            reels_df = pd.DataFrame(columns=[
                'reel_id', 'article_id', 'title', 'script_markdown',
                'creation_date', 'status', 'video_url', 'notes'
            ])
            
            logs_df = pd.DataFrame(columns=[
                'log_id', 'article_id', 'content_id', 'log_type',
                'timestamp', 'message', 'details'
            ])
            
            # Создаем Excel-файл с несколькими листами
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                content_df.to_excel(writer, sheet_name='Контент', index=False)
                articles_df.to_excel(writer, sheet_name='Статьи', index=False)
                schedule_df.to_excel(writer, sheet_name='Планирование', index=False)
                metadata_df.to_excel(writer, sheet_name='Метаданные', index=False)
                reels_df.to_excel(writer, sheet_name='Reels', index=False)
                logs_df.to_excel(writer, sheet_name='Логи', index=False)
                
            # Применяем форматирование
            self.apply_formatting()
    
    def apply_formatting(self):
        """Применяет форматирование к Excel-файлу"""
        wb = openpyxl.load_workbook(self.excel_path)
        
        # Форматирование для всех листов
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Заголовки жирным шрифтом
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Автофильтр
            ws.auto_filter.ref = ws.dimensions
            
            # Устанавливаем ширину столбцов для текстовых полей
            for col in ws.columns:
                if col[0].value in ['content_markdown', 'article_content', 'script_markdown', 'message', 'details']:
                    ws.column_dimensions[col[0].column_letter].width = 80
                else:
                    ws.column_dimensions[col[0].column_letter].width = 20
        
        # Сохраняем изменения
        wb.save(self.excel_path)
    
    def add_article(self, article_data, article_content=None):
        """Добавляет новую статью в таблицу"""
        # Генерируем уникальный ID
        article_id = str(uuid.uuid4())
        
        # Загружаем текущие данные
        articles_df = pd.read_excel(self.excel_path, sheet_name='Статьи')
        
        # Если передан контент статьи в виде словаря, преобразуем его в JSON строку
        if article_content and isinstance(article_content, dict):
            article_content = json.dumps(article_content, ensure_ascii=False)
        
        # Если передан контент статьи в виде списка строк, объединяем их
        if article_content and isinstance(article_content, list):
            article_content = "\n\n".join(article_content)
        
        # Создаем новую запись
        new_article = {
            'article_id': article_id,
            'title': article_data.get('title', ''),
            'source_url': article_data.get('link', ''),
            'category': article_data.get('category', ''),
            'publication_date': article_data.get('date', ''),
            'company_name': article_data.get('company_name', ''),
            'funding_amount': article_data.get('funding_amount', ''),
            'article_content': article_content or article_data.get('content', ''),
            'processing_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'processing_status': 'processed'
        }
        
        # Добавляем запись в DataFrame
        articles_df = pd.concat([articles_df, pd.DataFrame([new_article])], ignore_index=True)
        
        # Сохраняем обновленные данные
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            articles_df.to_excel(writer, sheet_name='Статьи', index=False)
        
        # Обновляем метаданные
        self.update_metadata('last_update', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Добавляем запись в лог
        self.add_log(article_id=article_id, log_type='article_added', 
                    message=f"Добавлена новая статья: {article_data.get('title', '')}")
        
        return article_id
    
    def add_content(self, content_data, article_id, content_markdown=None):
        """Добавляет новый контент, связанный со статьей"""
        # Генерируем уникальный ID
        content_id = str(uuid.uuid4())
        
        # Загружаем текущие данные
        content_df = pd.read_excel(self.excel_path, sheet_name='Контент')
        
        # Если контент передан в виде пути к файлу, читаем его содержимое
        if not content_markdown and 'content_path' in content_data and os.path.exists(content_data['content_path']):
            try:
                with open(content_data['content_path'], 'r', encoding='utf-8') as f:
                    content_markdown = f.read()
            except Exception as e:
                print(f"Ошибка при чтении файла контента: {e}")
        
        # Создаем новую запись
        new_content = {
            'content_id': content_id,
            'article_id': article_id,
            'title': content_data.get('title', ''),
            'source_url': content_data.get('source_url', ''),
            'category': content_data.get('category', ''),
            'content_type': content_data.get('content_type', ''),
            'language': content_data.get('language', ''),
            'content_markdown': content_markdown or '',
            'creation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'status': 'draft',
            'scheduled_date': '',
            'scheduled_time': '',
            'platform': content_data.get('platform', ''),
            'published_date': '',
            'published_url': '',
            'engagement_stats': '',
            'tags': content_data.get('tags', ''),
            'dubskiy_rating': content_data.get('dubskiy_rating', ''),
            'notes': content_data.get('notes', '')
        }
        
        # Добавляем запись в DataFrame
        content_df = pd.concat([content_df, pd.DataFrame([new_content])], ignore_index=True)
        
        # Сохраняем обновленные данные
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            content_df.to_excel(writer, sheet_name='Контент', index=False)
        
        # Обновляем метаданные
        self.update_metadata('last_update', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Добавляем запись в лог
        self.add_log(article_id=article_id, content_id=content_id, log_type='content_added', 
                    message=f"Добавлен новый контент типа {content_data.get('content_type', '')} на языке {content_data.get('language', '')}")
        
        return content_id
    
    def add_reel(self, article_id, title, script_markdown, notes=''):
        """Добавляет новый скрипт для Instagram Reel"""
        # Генерируем уникальный ID
        reel_id = str(uuid.uuid4())
        
        # Загружаем текущие данные
        reels_df = pd.read_excel(self.excel_path, sheet_name='Reels')
        
        # Создаем новую запись
        new_reel = {
            'reel_id': reel_id,
            'article_id': article_id,
            'title': title,
            'script_markdown': script_markdown,
            'creation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'status': 'draft',
            'video_url': '',
            'notes': notes
        }
        
        # Добавляем запись в DataFrame
        reels_df = pd.concat([reels_df, pd.DataFrame([new_reel])], ignore_index=True)
        
        # Сохраняем обновленные данные
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            reels_df.to_excel(writer, sheet_name='Reels', index=False)
        
        # Обновляем метаданные
        self.update_metadata('last_update', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Добавляем запись в лог
        self.add_log(article_id=article_id, log_type='reel_added', 
                    message=f"Добавлен новый скрипт для Instagram Reel: {title}")
        
        return reel_id
    
    def add_log(self, log_type, message, article_id=None, content_id=None, details=None):
        """Добавляет новую запись в лог"""
        # Генерируем уникальный ID
        log_id = str(uuid.uuid4())
        
        # Загружаем текущие данные
        logs_df = pd.read_excel(self.excel_path, sheet_name='Логи')
        
        # Создаем новую запись
        new_log = {
            'log_id': log_id,
            'article_id': article_id or '',
            'content_id': content_id or '',
            'log_type': log_type,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'message': message,
            'details': json.dumps(details) if details else ''
        }
        
        # Добавляем запись в DataFrame
        logs_df = pd.concat([logs_df, pd.DataFrame([new_log])], ignore_index=True)
        
        # Сохраняем обновленные данные
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            logs_df.to_excel(writer, sheet_name='Логи', index=False)
        
        return log_id
    
    def schedule_content(self, content_id, schedule_data):
        """Планирует публикацию контента"""
        # Генерируем уникальный ID для расписания
        schedule_id = str(uuid.uuid4())
        
        # Загружаем текущие данные
        schedule_df = pd.read_excel(self.excel_path, sheet_name='Планирование')
        content_df = pd.read_excel(self.excel_path, sheet_name='Контент')
        
        # Создаем новую запись в расписании
        new_schedule = {
            'schedule_id': schedule_id,
            'content_id': content_id,
            'platform': schedule_data.get('platform', ''),
            'scheduled_date': schedule_data.get('date', ''),
            'scheduled_time': schedule_data.get('time', ''),
            'timezone': schedule_data.get('timezone', 'UTC+3'),
            'posting_status': 'pending',
            'priority': schedule_data.get('priority', 'medium'),
            'campaign_id': schedule_data.get('campaign_id', ''),
            'posting_account': schedule_data.get('account', '')
        }
        
        # Добавляем запись в DataFrame расписания
        schedule_df = pd.concat([schedule_df, pd.DataFrame([new_schedule])], ignore_index=True)
        
        # Обновляем статус контента
        content_mask = content_df['content_id'] == content_id
        if any(content_mask):
            content_df.loc[content_mask, 'status'] = 'scheduled'
            content_df.loc[content_mask, 'scheduled_date'] = schedule_data.get('date', '')
            content_df.loc[content_mask, 'scheduled_time'] = schedule_data.get('time', '')
        
        # Сохраняем обновленные данные
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            schedule_df.to_excel(writer, sheet_name='Планирование', index=False)
            content_df.to_excel(writer, sheet_name='Контент', index=False)
        
        # Обновляем метаданные
        self.update_metadata('last_update', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Добавляем запись в лог
        article_id = ''
        if any(content_mask):
            article_id = content_df.loc[content_mask, 'article_id'].values[0]
        
        self.add_log(article_id=article_id, content_id=content_id, log_type='content_scheduled', 
                    message=f"Запланирована публикация контента на {schedule_data.get('date', '')} {schedule_data.get('time', '')}")
        
        return schedule_id
    
    def update_metadata(self, key, value):
        """Обновляет значение в метаданных"""
        metadata_df = pd.read_excel(self.excel_path, sheet_name='Метаданные')
        
        # Обновляем значение
        key_mask = metadata_df['key'] == key
        if any(key_mask):
            metadata_df.loc[key_mask, 'value'] = value
        else:
            # Добавляем новую запись, если ключ не существует
            metadata_df = pd.concat([metadata_df, pd.DataFrame([{'key': key, 'value': value}])], ignore_index=True)
        
        # Сохраняем обновленные данные
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            metadata_df.to_excel(writer, sheet_name='Метаданные', index=False)
    
    def get_all_articles(self):
        """Возвращает все статьи из таблицы"""
        try:
            articles_df = pd.read_excel(self.excel_path, sheet_name='Статьи')
            return articles_df.to_dict(orient='records')
        except Exception as e:
            print(f"Ошибка при чтении статей: {e}")
            return []
    
    def get_all_content(self):
        """Возвращает весь контент из таблицы"""
        try:
            content_df = pd.read_excel(self.excel_path, sheet_name='Контент')
            return content_df.to_dict(orient='records')
        except Exception as e:
            print(f"Ошибка при чтении контента: {e}")
            return []
    
    def get_content_by_id(self, content_id):
        """Возвращает контент по его ID"""
        try:
            content_df = pd.read_excel(self.excel_path, sheet_name='Контент')
            content_mask = content_df['content_id'] == content_id
            if any(content_mask):
                return content_df[content_mask].to_dict(orient='records')[0]
            else:
                print(f"Контент с ID {content_id} не найден")
                return None
        except Exception as e:
            print(f"Ошибка при чтении контента: {e}")
            return None
    
    def get_article_by_id(self, article_id):
        """Возвращает статью по её ID"""
        try:
            articles_df = pd.read_excel(self.excel_path, sheet_name='Статьи')
            article_mask = articles_df['article_id'] == article_id
            if any(article_mask):
                return articles_df[article_mask].to_dict(orient='records')[0]
            else:
                print(f"Статья с ID {article_id} не найдена")
                return None
        except Exception as e:
            print(f"Ошибка при чтении статьи: {e}")
            return None
    
    def get_reels_by_article_id(self, article_id):
        """Возвращает все reels для указанной статьи"""
        try:
            reels_df = pd.read_excel(self.excel_path, sheet_name='Reels')
            reels_mask = reels_df['article_id'] == article_id
            if any(reels_mask):
                return reels_df[reels_mask].to_dict(orient='records')
            else:
                return []
        except Exception as e:
            print(f"Ошибка при чтении reels: {e}")
            return []
    
    def get_logs_by_article_id(self, article_id):
        """Возвращает все логи для указанной статьи"""
        try:
            logs_df = pd.read_excel(self.excel_path, sheet_name='Логи')
            logs_mask = logs_df['article_id'] == article_id
            if any(logs_mask):
                return logs_df[logs_mask].to_dict(orient='records')
            else:
                return []
        except Exception as e:
            print(f"Ошибка при чтении логов: {e}")
            return []
    
    def export_content_to_file(self, content_id, output_path):
        """Экспортирует контент в файл"""
        content = self.get_content_by_id(content_id)
        if content and content.get('content_markdown'):
            try:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(content['content_markdown'])
                return True
            except Exception as e:
                print(f"Ошибка при экспорте контента: {e}")
                return False
        return False
    
    def import_content_from_file(self, content_id, input_path):
        """Импортирует контент из файла"""
        if os.path.exists(input_path):
            try:
                with open(input_path, 'r', encoding='utf-8') as f:
                    content_markdown = f.read()
                
                content_df = pd.read_excel(self.excel_path, sheet_name='Контент')
                content_mask = content_df['content_id'] == content_id
                
                if any(content_mask):
                    content_df.loc[content_mask, 'content_markdown'] = content_markdown
                    
                    with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        content_df.to_excel(writer, sheet_name='Контент', index=False)
                    
                    return True
            except Exception as e:
                print(f"Ошибка при импорте контента: {e}")
        return False 